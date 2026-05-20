from openpyxl import load_workbook
from app.utils.exceptions import AppError

EXPECTED_HEADERS = [
    "SKU",
    "Nombre",
    "Stock Actual",
    "Stock Mínimo",
    "Precio Estandar",
    "Proveedor",
    "Descripción",
    "Estado",
]

def load_excel_rows(file):
    try: 
        workbook = load_workbook(file, data_only=True)
    except Exception:
        raise AppError(400, "Invalid Excel file")
    
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]

    if headers != EXPECTED_HEADERS:
        raise AppError(
            400, 
            f"Invalid Excel format. Expected headers: {EXPECTED_HEADERS}"
        )
    
    rows = []

    for row in sheet.iter_rows(min_row = 2, values_only = True):
        rows.append(row)

    return rows